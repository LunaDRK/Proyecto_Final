import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
import tkinter as tk
from tkinter import messagebox

def limpiar_ventana(ventana):
    for widget in ventana.winfo_children():
        widget.destroy()

def mostrar_clientes(ventana, volver_callback):
    limpiar_ventana(ventana)

    # Boton de regresar
    btn_volver = tk.Button(ventana, text="Regresar", width=25, height=2, bg="#E04E4E", fg="white",
                           command=volver_callback)
    btn_volver.pack(pady=10, side='top')

    # Nombre del archivo Excel
    ARCHIVO = "algoritmos/transacciones_ventas/proyecto.xlsx"

    # Crear archivo si no existe
    def crear_excel():
        if not os.path.exists(ARCHIVO):
            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Clientes"
            hoja["A1"] = "Código"
            hoja["B1"] = "Nombre"
            hoja["C1"] = "Dirección"
            wb.save(ARCHIVO)

    # Leer datos del Excel
    def leer_clientes():
        wb = openpyxl.load_workbook(ARCHIVO)
        hoja = wb.active
        datos = []
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            datos.append(fila)
        wb.close()
        return datos

    # Escribir nuevo cliente
    def escribir_cliente(codigo, nombre, direccion):
        if not codigo or not nombre or not direccion:
            messagebox.showwarning("Campos vacíos", "Debes llenar todos los campos.")
            return

        wb = openpyxl.load_workbook(ARCHIVO)
        hoja = wb.active
        hoja.append([codigo, nombre, direccion])
        wb.save(ARCHIVO)
        wb.close()
        messagebox.showinfo("Éxito", "Cliente agregado correctamente.")
        actualizar_tabla()  # Refresca la tabla
        limpiar_campos()

    # Actualizar tabla completa
    def actualizar_tabla():
        for fila in tabla.get_children():
            tabla.delete(fila)
        for cliente in leer_clientes():
            tabla.insert("", "end", values=cliente)

    # Buscar clientes según criterio
    def buscar_cliente():
        texto = entry_buscar.get().strip().lower()
        if not texto:
            messagebox.showinfo("Buscar", "Por favor ingrese algo para buscar.")
            return

        resultados = []
        # Buscar en todos los campos de cada cliente
        for cliente in leer_clientes():
            if any(texto in str(campo).lower() for campo in cliente if campo):
                resultados.append(cliente)

        # Mostrar resultados
        for fila in tabla.get_children():
            tabla.delete(fila)

        if resultados:
            for cliente in resultados:
                tabla.insert("", "end", values=cliente)
        else:
            messagebox.showinfo("Sin resultados", "No se encontraron coincidencias.")

    # Limpiar campos después de agregar
    def limpiar_campos():
        entry_codigo.delete(0, tk.END)
        entry_nombre.delete(0, tk.END)
        entry_direccion.delete(0, tk.END)


    #editar clientes
    def editar_clientes():
        cliente_seleccionado = tabla.focus()
        if not cliente_seleccionado:
            messagebox.showwarning("Editar", "Selecciona dando un click sobre un cliente para editar")
            return

        valores_originales = tabla.item(cliente_seleccionado, "values")
        codigo_original, nombre_original, direccion_original = valores_originales

        #crea una ventana emergente para un form de editar
        ventana_editar = tk.Toplevel(ventana)
        ventana_editar.title("Editar Cliente")
        ventana_editar.geometry("400x300")
        ventana_editar.configure(bg="#F0F0F0")

        tk.Label(ventana_editar, text="Código:", font=("Arial", 12), bg="#F0F0F0").grid(row=0, column=0, padx=10, pady=10)
        entry_codigo_editar = tk.Entry(ventana_editar, width=30)
        entry_codigo_editar.grid(row=0, column=1, padx=10, pady=10)
        entry_codigo_editar.insert(0, codigo_original)

        tk.Label(ventana_editar, text="Nombre:", font=("Arial", 12), bg="#F0F0F0").grid(row=1, column=0, padx=10, pady=10)
        entry_nombre_editar = tk.Entry(ventana_editar, width=30)
        entry_nombre_editar.grid(row=1, column=1, padx=10, pady=10)
        entry_nombre_editar.insert(0, nombre_original)

        tk.Label(ventana_editar, text="Dirección:", font=("Arial", 12), bg="#F0F0F0").grid(row=2, column=0, padx=10, pady=10)
        entry_direccion_editar = tk.Entry(ventana_editar, width=30)
        entry_direccion_editar.grid(row=2, column=1, padx=10, pady=10)
        entry_direccion_editar.insert(0, direccion_original)

        def guardar_cambios():
            codigo_editado = entry_codigo_editar.get().strip()
            nombre_editado = entry_nombre_editar.get().strip()
            direccion_editada = entry_direccion_editar.get().strip()

            if not codigo_editado or not nombre_editado or not direccion_editada:
                messagebox.showwarning("Campos vacíos", "Debes llenar todos los campos.")
                return

            # Actualiza el excel
            wb = openpyxl.load_workbook(ARCHIVO)
            hoja = wb.active
            for fila in hoja.iter_rows(min_row=2):
                if str(fila[0].value) == str(codigo_original):
                    fila[0].value = codigo_editado
                    fila[1].value = nombre_editado
                    fila[2].value = direccion_editada
                    break

            wb.save(ARCHIVO)
            wb.close()

            # Actualiza solo la fila que se editó y muestra mensaje que se editó correctamente
            tabla.item(cliente_seleccionado, values=(codigo_editado, nombre_editado, direccion_editada))
            messagebox.showinfo("Editado", "Información de Cliente actualizada correctamente")
            ventana_editar.destroy()  # Cierra la ventana de edición

        tk.Button(ventana_editar, text="Guardar Cambios", bg="#FFC107", fg="black",
                        font=("Arial", 12, "bold"), command=guardar_cambios).grid(row=3, column=0, columnspan=2, pady=20)

    #funcion de eliminar datos del cliente
    def eliminar_cliente():
        cliente_seleccionado = tabla.focus()
        if not cliente_seleccionado:
            messagebox.showwarning("Eliminar", "Selecciona un cliente para eliminar")
            return

        valores = tabla.item(cliente_seleccionado, "values")
        codigo_cliente = valores[0]
        
        confirmar = messagebox.askyesno("Confirmar Eliminación", f"¿Estás seguro de eliminar al cliente con código {codigo_cliente}?")
        if not confirmar:
            return

        # Elimina del excel
        wb = openpyxl.load_workbook(ARCHIVO)
        hoja = wb.active
        for fila in hoja.iter_rows(min_row=2):
            if str(fila[0].value) == str(codigo_cliente):
                hoja.delete_rows(fila[0].row, 1)
                break

        wb.save(ARCHIVO)
        wb.close()

        # Elimina de la tabla y muestra mensaje que se eliminó correctamente
        tabla.delete(cliente_seleccionado)
        messagebox.showinfo("Eliminado", "Cliente eliminado correctamente.")

    #Configuración de la ventana
    """ventana = tk.Tk()
    ventana.title("Gestión de Clientes")
    ventana.geometry("900x600")
    ventana.configure(bg="#F0F0F0")"""

    #boton de regresar
    """btn_regresar = tk.Button(ventana, text="Regresar", width=25, height=2, bg="#9E9E9E", fg="white", font=("Arial", 12, "bold"),
                                            command=lambda: regresar("Regresar"))
    btn_regresar.pack(pady=10)"""

    # Crear archivo si no existe
    crear_excel()

    titulo = tk.Label(ventana, text="Módulo de Clientes", font=("Arial", 18, "bold"), bg="#F0F0F0")
    titulo.pack(pady=15, side='top')

    # Frame de formulario
    frame_form = tk.Frame(ventana, bg="#F0F0F0")
    frame_form.pack(pady=10)

    tk.Label(frame_form, text="Código:", font=("Arial", 12), bg="#F0F0F0").grid(row=0, column=0, padx=10, pady=5)
    entry_codigo = tk.Entry(frame_form, width=30)
    entry_codigo.grid(row=0, column=1, padx=10, pady=5)

    tk.Label(frame_form, text="Nombre:", font=("Arial", 12), bg="#F0F0F0").grid(row=1, column=0, padx=10, pady=5)
    entry_nombre = tk.Entry(frame_form, width=30)
    entry_nombre.grid(row=1, column=1, padx=10, pady=5)

    tk.Label(frame_form, text="Dirección:", font=("Arial", 12), bg="#F0F0F0").grid(row=2, column=0, padx=10, pady=5)
    entry_direccion = tk.Entry(frame_form, width=30)
    entry_direccion.grid(row=2, column=1, padx=10, pady=5)

    # Botón de agregar
    btn_agregar = tk.Button(ventana, text="Agregar Cliente", bg="#4CAF50", fg="white",
                            font=("Arial", 12, "bold"), width=20,
                            command=lambda: escribir_cliente(entry_codigo.get(), entry_nombre.get(), entry_direccion.get()))
    btn_agregar.pack(pady=10)

    # Frame de búsqueda
    frame_buscar = tk.Frame(ventana, bg="#F0F0F0")
    frame_buscar.pack(pady=10)

    tk.Label(frame_buscar, text="Buscar:", font=("Arial", 12), bg="#F0F0F0").grid(row=0, column=0, padx=10)
    entry_buscar = tk.Entry(frame_buscar, width=40)
    entry_buscar.grid(row=0, column=1, padx=10)
    btn_buscar = tk.Button(frame_buscar, text="Buscar", bg="#2196F3", fg="white",
                        font=("Arial", 11, "bold"), command=buscar_cliente)
    btn_buscar.grid(row=0, column=2, padx=10)

    btn_todos = tk.Button(frame_buscar, text="Mostrar Todos", bg="#9E9E9E", fg="white",
                        font=("Arial", 11, "bold"), command=actualizar_tabla)
    btn_todos.grid(row=0, column=3, padx=10)

    #btn de editar y eliminar
    frame_botones = tk.Frame(ventana, bg="#F0F0F0")
    frame_botones.pack(pady=10)

    tk.Button(frame_buscar, text="Editar", bg="#FFC107", fg="black", font=("Arial", 12, "bold"),
                    command=editar_clientes).grid(row=0, column=4, padx=5)

    tk.Button(frame_buscar, text="Eliminar", bg="#F44336", fg="white",
            font=("Arial", 12, "bold"), command=eliminar_cliente).grid(row=0, column=5, padx=5)


    # Tabla de clientes
    frame_tabla = tk.Frame(ventana)
    frame_tabla.pack(pady=10)

    columnas = ("Código", "Nombre", "Dirección")
    tabla = ttk.Treeview(frame_tabla, columns=columnas, show="headings", height=12)
    for col in columnas:
        tabla.heading(col, text=col)
        tabla.column(col, width=200, anchor="center")
    tabla.pack()

    # Cargar datos iniciales
    actualizar_tabla()

    #ventana.mainloop()
