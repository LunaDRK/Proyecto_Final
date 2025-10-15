import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os

def limpiar_ventana(ventana):
    for widget in ventana.winfo_children():
        widget.destroy()

def mostrar_inventario(ventana, volver_callback):
    limpiar_ventana(ventana)

    # Botón de regresar
    btn_volver = tk.Button(ventana, text="Regresar", width=25, height=2, bg="#E04E4E", fg="white",
                           command=volver_callback)
    btn_volver.pack(pady=10, side='top')

    #Base de datos y tabla de inventario
    ARCHIVO = "algoritmos/transacciones_ventas/proyecto.xlsx"
    HOJA = "Inventario"

    # Crear hoja si no existen 
    def crear_excel_y_hoja():
        if not os.path.exists(ARCHIVO):
            wb = openpyxl.Workbook()
            if HOJA in wb.sheetnames:
                hoja = wb[HOJA]
            else:
                hoja = wb.active
                hoja.title = HOJA
            hoja["A1"] = "Código"
            hoja["B1"] = "Nombre"
            hoja["C1"] = "Existencia"
            hoja["D1"] = "Proveedor"
            hoja["E1"] = "Precio"
            wb.save(ARCHIVO)
            wb.close()
        else:
            wb = openpyxl.load_workbook(ARCHIVO)
            if HOJA not in wb.sheetnames:
                hoja = wb.create_sheet(HOJA)
                hoja["A1"] = "Código"
                hoja["B1"] = "Nombre"
                hoja["C1"] = "Existencia"
                hoja["D1"] = "Proveedor"
                hoja["E1"] = "Precio"
                wb.save(ARCHIVO)
            wb.close()

    # Leer productos 
    def leer_productos():
        wb = openpyxl.load_workbook(ARCHIVO)
        hoja = wb[HOJA]
        datos = []
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            datos.append(fila)
        wb.close()
        return datos

    # Escribir nuevo producto 
    def escribir_producto(codigo, nombre, existencia, proveedor, precio):
        if not codigo or not nombre or not existencia or not proveedor or not precio:
            messagebox.showwarning("Campos vacíos", "Debes llenar todos los campos.")
            return

        try:
            existencia = int(existencia)
            precio = float(precio)
        except ValueError:
            messagebox.showerror("Error", "Existencia debe ser entero y Precio número (ej: 12.50).")
            return

        # Append en la hoja correcta sin tocar otras hojas
        wb = openpyxl.load_workbook(ARCHIVO)
        hoja = wb[HOJA]
        hoja.append([codigo, nombre, existencia, proveedor, precio])
        wb.save(ARCHIVO)
        wb.close()
        messagebox.showinfo("Éxito", "Producto agregado correctamente.")
        actualizar_tabla()
        limpiar_campos()

    # Actualizar tabla en la interfaz
    def actualizar_tabla():
        for fila in tabla.get_children():
            tabla.delete(fila)
        for producto in leer_productos():
            tabla.insert("", "end", values=producto)

    # Buscar productos 
    def buscar_producto():
        texto = entry_buscar.get().strip().lower()
        if not texto:
            messagebox.showinfo("Buscar", "Ingrese algo para buscar.")
            return

        resultados = []
        for producto in leer_productos():
            if any(texto in str(campo).lower() for campo in producto if campo is not None):
                resultados.append(producto)

        for fila in tabla.get_children():
            tabla.delete(fila)

        if resultados:
            for producto in resultados:
                tabla.insert("", "end", values=producto)
        else:
            messagebox.showinfo("Sin resultados", "No se encontraron coincidencias.")

    def limpiar_campos():
        entry_codigo.delete(0, tk.END)
        entry_nombre.delete(0, tk.END)
        entry_existencia.delete(0, tk.END)
        entry_proveedor.delete(0, tk.END)
        entry_precio.delete(0, tk.END)

    # Inicializador
    crear_excel_y_hoja()

    titulo = tk.Label(ventana, text="Control de Inventario", font=("Arial", 18, "bold"), bg="#F0F0F0")
    titulo.pack(pady=15, side='top')

    # Frame del formulario
    frame_form = tk.Frame(ventana, bg="#F0F0F0")
    frame_form.pack(pady=10)

    tk.Label(frame_form, text="Código:", font=("Arial", 12), bg="#F0F0F0").grid(row=0, column=0, padx=10, pady=5)
    entry_codigo = tk.Entry(frame_form, width=30)
    entry_codigo.grid(row=0, column=1, padx=10, pady=5)

    tk.Label(frame_form, text="Nombre:", font=("Arial", 12), bg="#F0F0F0").grid(row=1, column=0, padx=10, pady=5)
    entry_nombre = tk.Entry(frame_form, width=30)
    entry_nombre.grid(row=1, column=1, padx=10, pady=5)

    tk.Label(frame_form, text="Existencia:", font=("Arial", 12), bg="#F0F0F0").grid(row=2, column=0, padx=10, pady=5)
    entry_existencia = tk.Entry(frame_form, width=30)
    entry_existencia.grid(row=2, column=1, padx=10, pady=5)

    tk.Label(frame_form, text="Proveedor:", font=("Arial", 12), bg="#F0F0F0").grid(row=3, column=0, padx=10, pady=5)
    entry_proveedor = tk.Entry(frame_form, width=30)
    entry_proveedor.grid(row=3, column=1, padx=10, pady=5)

    tk.Label(frame_form, text="Precio:", font=("Arial", 12), bg="#F0F0F0").grid(row=4, column=0, padx=10, pady=5)
    entry_precio = tk.Entry(frame_form, width=30)
    entry_precio.grid(row=4, column=1, padx=10, pady=5)

    btn_agregar = tk.Button(ventana, text="Agregar Producto", bg="#4CAF50", fg="white",
                            font=("Arial", 12, "bold"), width=20,
                            command=lambda: escribir_producto(entry_codigo.get(), entry_nombre.get(),
                                                              entry_existencia.get(), entry_proveedor.get(),
                                                              entry_precio.get()))
    btn_agregar.pack(pady=10)

    # Frame de búsqueda
    frame_buscar = tk.Frame(ventana, bg="#F0F0F0")
    frame_buscar.pack(pady=10)

    tk.Label(frame_buscar, text="Buscar:", font=("Arial", 12), bg="#F0F0F0").grid(row=0, column=0, padx=10)
    entry_buscar = tk.Entry(frame_buscar, width=40)
    entry_buscar.grid(row=0, column=1, padx=10)
    btn_buscar = tk.Button(frame_buscar, text="Buscar", bg="#2196F3", fg="white",
                        font=("Arial", 11, "bold"), command=buscar_producto)
    btn_buscar.grid(row=0, column=2, padx=10)

    btn_todos = tk.Button(frame_buscar, text="Mostrar Todos", bg="#9E9E9E", fg="white",
                        font=("Arial", 11, "bold"), command=actualizar_tabla)
    btn_todos.grid(row=0, column=3, padx=10)

    # Tabla
    frame_tabla = tk.Frame(ventana)
    frame_tabla.pack(pady=10)

    columnas = ("Código", "Nombre", "Existencia", "Proveedor", "Precio")
    tabla = ttk.Treeview(frame_tabla, columns=columnas, show="headings", height=12)
    for col in columnas:
        tabla.heading(col, text=col)
        tabla.column(col, width=150, anchor="center")
    tabla.pack()

    actualizar_tabla()
