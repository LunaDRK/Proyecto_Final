import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
import ttkbootstrap as ttk
from ttkbootstrap.constants import *

def limpiar_ventana(ventana):
    for widget in ventana.winfo_children():
        widget.destroy()

def mostrar_ventas(ventana, volver_callback):
    limpiar_ventana(ventana)

     # Botón de regresar
    btn_volver = ttk.Button(ventana, text="← Regresar", width=15, padding=10, bootstyle="danger-outline",
                           command=volver_callback)
    btn_volver.pack(side="top", anchor="nw", padx=15, pady=15)

    ARCHIVO = "algoritmos/transacciones_ventas/proyecto.xlsx"

    # Creacion de DB y Tabla si no existe
    def crear_excel():
        if not os.path.exists(ARCHIVO):
            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Ventas"
            hoja.append(["Código Producto", "Código Cliente", "Cantidad", "Total"])
            wb.create_sheet("Clientes")
            wb.create_sheet("Inventario")
            wb.save(ARCHIVO)
        else:
            wb = openpyxl.load_workbook(ARCHIVO)
            for hoja_nombre in ["Ventas", "Clientes", "Inventario"]:
                if hoja_nombre not in wb.sheetnames:
                    hoja = wb.create_sheet(hoja_nombre)
                    if hoja_nombre == "Ventas":
                        hoja.append(["Código Producto", "Código Cliente", "Cantidad", "Total"])
            wb.save(ARCHIVO)
            wb.close()

    crear_excel()

    #Lectura de Tabla de la db
    def leer_hoja(nombre):
        wb = openpyxl.load_workbook(ARCHIVO)
        if nombre not in wb.sheetnames:
            wb.close()
            return []
        hoja = wb[nombre]
        datos = []
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            datos.append(fila)
        wb.close()
        return datos

    #Busqueda en la tabla de clientes
    def buscar_codigo_cliente(nombre):
        for codigo, nombre_cliente, *_ in leer_hoja("Clientes"):
            if nombre_cliente and nombre_cliente.lower() == nombre.lower():
                return codigo
        return None

    #Busqueda en la tabla de inventario
    def buscar_producto(nombre):
        for row in leer_hoja("Inventario"):
            # Se usa el 5 asumiendo el orden (codigo, nombre, existencia, proveedor, precio)
            if len(row) >= 5:
                codigo, nombre_producto, existencia, *_resto, precio = row
            else:
                # por si la hoja tiene menos columnas
                codigo = row[0] if len(row) > 0 else None
                nombre_producto = row[1] if len(row) > 1 else None
                existencia = row[2] if len(row) > 2 else 0
                precio = row[4] if len(row) > 4 else (row[-1] if row else 0)
            if nombre_producto and nombre_producto.lower() == nombre.lower():
                try:
                    return codigo, float(existencia), float(precio)
                except:
                    # en caso existencias/precio vacíos
                    return codigo, 0.0, 0.0
        return None, None, None
    
    #Buscar nombre del cliente por su codigo
    def buscar_nombre_cliente(codigo):
        for cod, nombre, *_ in leer_hoja("Clientes"):
            if cod == codigo:
                return nombre
        return f"Cliente {codigo}"
    
    #Buscar nombre del producto por su codigo
    def buscar_nombre_producto(codigo):
        for cod, nombre, *_ in leer_hoja("Inventario"):
            if cod == codigo:
                return nombre
        return f"Producto {codigo}"

    #Actualizar stock de producto en el inventario
    def actualizar_stock(codigo_producto, cantidad, sumar=False):
        wb = openpyxl.load_workbook(ARCHIVO)
        hoja = wb["Inventario"]
        for fila in hoja.iter_rows(min_row=2):
            if fila[0].value == codigo_producto:
                try:
                    actual = float(fila[2].value)
                except:
                    actual = 0.0
                fila[2].value = actual + cantidad if sumar else actual - cantidad
                break
        wb.save(ARCHIVO)
        wb.close()

    #Lista temporal de productos para la venta 
    productos_en_venta = []

    #Agregar a la lista temporal
    def agregar_producto():
        nombre_producto = entry_producto.get().strip()
        cantidad = entry_cantidad.get().strip()

        if not nombre_producto or not cantidad:
            messagebox.showwarning("Campos vacíos", "Debes llenar todos los campos de producto y cantidad.")
            return

        try:
            cantidad = int(cantidad)
            if cantidad <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Error", "La cantidad debe ser un número entero positivo.")
            return

        codigo_producto, existencia, precio = buscar_producto(nombre_producto)
        if not codigo_producto:
            messagebox.showerror("Error", f"No se encontró el producto '{nombre_producto}'.")
            return

        if existencia < cantidad:
            messagebox.showerror("Error", f"Stock insuficiente. Solo hay {existencia} unidades disponibles.")
            return

        total = cantidad * precio

        # Agregar a la lista temporal
        productos_en_venta.append({
            "codigo_producto": codigo_producto,
            "nombre_producto": nombre_producto,
            "cantidad": cantidad,
            "precio": precio,
            "total": total
        })

        # Actualizar tabla temporal
        actualizar_tabla_temporal()
        entry_producto.delete(0, tk.END)
        entry_cantidad.delete(0, tk.END)
        entry_producto.focus_set()

    # Muestra la tabla temporal
    def actualizar_tabla_temporal():
        for fila in tabla_temporal.get_children():
            tabla_temporal.delete(fila)
        for item in productos_en_venta:
            tabla_temporal.insert("", "end", values=(
                item["nombre_producto"], item["cantidad"], f"Q{item['precio']:.2f}", f"Q{item['total']:.2f}"
            ))
        actualizar_total_carrito()

    #Muestra el total del carrito
    def actualizar_total_carrito():
        total = sum(item["total"] for item in productos_en_venta)
        label_total_var.set(f"Total carrito: Q{total:.2f}")

    #Gurda la venta en la DB
    def guardar_venta():
        nombre_cliente = entry_cliente.get().strip()
        if not nombre_cliente:
            messagebox.showwarning("Campos vacíos", "Debes ingresar el nombre del cliente.")
            return

        if not productos_en_venta:
            messagebox.showwarning("Sin productos", "Agrega al menos un producto antes de registrar la venta.")
            return

        codigo_cliente = buscar_codigo_cliente(nombre_cliente)
        if not codigo_cliente:
            messagebox.showerror("Error", f"No se encontró el cliente '{nombre_cliente}'.")
            return

        # Validar existencias antes de guardar
        for item in productos_en_venta:
            _, existencia, _ = buscar_producto(item["nombre_producto"])
            if existencia < item["cantidad"]:
                messagebox.showerror("Error", f"El producto '{item['nombre_producto']}' no tiene suficiente stock.")
                return

        # Registrar venta (una fila por producto)
        wb = openpyxl.load_workbook(ARCHIVO)
        hoja_ventas = wb["Ventas"]
        hoja_inventario = wb["Inventario"]

        for item in productos_en_venta:
            # Agregar a la hoja de ventas
            hoja_ventas.append([
                item["codigo_producto"],
                codigo_cliente,
                item["cantidad"],
                item["total"]
            ])

    # Actualizar stock directamente
        for fila in hoja_inventario.iter_rows(min_row=2):
            if fila[0].value == item["codigo_producto"]:
                try:
                    actual = float(fila[2].value)
                except:
                    actual = 0.0
                fila[2].value = actual - item["cantidad"]
                break

        wb.save(ARCHIVO)
        wb.close()

        messagebox.showinfo("Éxito", "Venta registrada correctamente.")
        productos_en_venta.clear()
        actualizar_tabla_temporal()
        actualizar_tabla()
        entry_cliente.delete(0, tk.END)

    #Anula la venta seleccionada
    def anular_venta():
        seleccion = tabla.focus()
        if not seleccion:
            messagebox.showwarning("Eliminar", "Selecciona una venta para anular.")
            return

        valores = tabla.item(seleccion, "values")
        if not valores:
            return

        producto_nombre, cliente_nombre, cantidad, total = valores
        cantidad = int(cantidad)
        codigo_cliente = buscar_codigo_cliente(cliente_nombre)
        codigo_producto, _, _ = buscar_producto(producto_nombre)

        confirmar = messagebox.askyesno("Confirmar", f"¿Deseas anular esta venta?\n{valores}")
        if not confirmar:
            return

        wb = openpyxl.load_workbook(ARCHIVO)
        hoja = wb["Ventas"]
        # Eliminar todas las coincidencias de esa venta
        filas_a_eliminar = []
        for fila in range(2, hoja.max_row + 1):
            if (str(hoja.cell(fila, 1).value) == str(codigo_producto) and
                str(hoja.cell(fila, 2).value) == str(codigo_cliente) and
                int(hoja.cell(fila, 3).value) == cantidad):
                filas_a_eliminar.append(fila)

        for fila in reversed(filas_a_eliminar):
            hoja.delete_rows(fila, 1)

        wb.save(ARCHIVO)
        wb.close()

        actualizar_stock(codigo_producto, cantidad, sumar=True)
        tabla.delete(seleccion)
        messagebox.showinfo("Eliminado", "Venta anulada correctamente.")

    # Actualiza la tabla de ventas
    def actualizar_tabla(filtro=""):
        for fila in tabla.get_children():
            tabla.delete(fila)
        ventas = leer_hoja("Ventas")
        for codigo_producto, codigo_cliente, cantidad, total in ventas:
            nombre_producto = buscar_nombre_producto(codigo_producto)
            nombre_cliente = buscar_nombre_cliente(codigo_cliente)
            if filtro.lower() in nombre_cliente.lower() or filtro.lower() in nombre_producto.lower():
                tabla.insert("", "end", values=(nombre_producto, nombre_cliente, cantidad, total))

    # Autocompletado para entradas de nombre y producto
    def autocompletar(entry, lista):
        ventana_autocomplete = tk.Listbox(ventana, height=5, width=30)

        def actualizar_lista(event):
            texto = entry.get().lower()
            coincidencias = [item for item in lista if texto in item.lower()]
            ventana_autocomplete.delete(0, tk.END)
            for c in coincidencias[:8]:
                ventana_autocomplete.insert(tk.END, c)
            if coincidencias:
                x = entry.winfo_rootx() - ventana.winfo_rootx()
                y = entry.winfo_rooty() - ventana.winfo_rooty() + entry.winfo_height()
                ventana_autocomplete.place(x=x, y=y)
            else:
                ventana_autocomplete.place_forget()

        def seleccionar_item(event):
            if not ventana_autocomplete.curselection():
                return
            seleccionado = ventana_autocomplete.get(ventana_autocomplete.curselection())
            entry.delete(0, tk.END)
            entry.insert(0, seleccionado)
            ventana_autocomplete.place_forget()

        entry.bind("<KeyRelease>", actualizar_lista)
        ventana_autocomplete.bind("<<ListboxSelect>>", seleccionar_item)

    #Interfaz del modulo con 2 columnas
    titulo = tk.Label(ventana, text="Módulo de Ventas", font=("Arial", 18, "bold"), bg="#F0F0F0")
    titulo.pack(pady=5, side='top')

    frame_principal = tk.Frame(ventana, bg="#F0F0F0")
    frame_principal.pack(fill="both", expand=True, padx=10, pady=5)

    # Izquierda crear venta y tabla temporal
    frame_izq = tk.Frame(frame_principal, bg="#F0F0F0")
    frame_izq.pack(side="left", fill="both", expand=True, padx=5, pady=5)

    # Derecha ventas realizadas buscador y anular
    frame_der = tk.Frame(frame_principal, bg="#F0F0F0")
    frame_der.pack(side="right", fill="both", expand=True, padx=5, pady=5)

    #Modulo izquierdo
    tk.Label(frame_izq, text="Crear venta", font=("Arial", 14, "bold"), bg="#F0F0F0").pack(pady=8)

    # Frame del formulario centrado
    frame_form = tk.Frame(frame_izq, bg="#F0F0F0")
    frame_form.pack(pady=10)

    # Configurar columnas para que se centren
    frame_form.columnconfigure(0, weight=1)
    frame_form.columnconfigure(1, weight=1)

    # Campos del formulario centrados
    tk.Label(frame_form, text="Cliente:", font=("Arial", 12), bg="#F0F0F0").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    entry_cliente = tk.Entry(frame_form, width=30)
    entry_cliente.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    tk.Label(frame_form, text="Producto:", font=("Arial", 12), bg="#F0F0F0").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    entry_producto = tk.Entry(frame_form, width=30)
    entry_producto.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    tk.Label(frame_form, text="Cantidad:", font=("Arial", 12), bg="#F0F0F0").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    entry_cantidad = tk.Entry(frame_form, width=30)
    entry_cantidad.grid(row=2, column=1, padx=5, pady=5, sticky="w")

    # Frame contenedor de los botones
    frame_botones = ttk.Frame(frame_izq)
    frame_botones.pack(pady=10)  # espacio vertical

    # Botones uno al lado del otro
    ttk.Button(frame_botones, text="Agregar Producto", padding=10, bootstyle="success-outline",
           command=agregar_producto, width=20).pack(side=tk.LEFT, padx=10)

    ttk.Button(frame_botones, text="Registrar Venta", padding=10, bootstyle="warning-outline",
           width=20, command=guardar_venta).pack(side=tk.LEFT, padx=10)

    # Tabla temporal en la izquierda
    tk.Label(frame_izq, text="Productos en la venta:", font=("Arial", 12), bg="#F0F0F0").pack(pady=5)
    frame_tabla_temp = tk.Frame(frame_izq, bg="#F0F0F0")
    frame_tabla_temp.pack(fill="both", expand=True, padx=5, pady=5)
    columnas_temp = ("Producto", "Cantidad", "Precio", "Total")
    tabla_temporal = ttk.Treeview(frame_tabla_temp, columns=columnas_temp, show="headings", height=8)
    for col in columnas_temp:
        tabla_temporal.heading(col, text=col)
        tabla_temporal.column(col, width=120, anchor="center")
    tabla_temporal.pack(fill="both", expand=True)

    # label total carrito
    label_total_var = tk.StringVar(value="Total carrito: Q0.00")
    tk.Label(frame_izq, textvariable=label_total_var, font=("Arial", 12, "bold"), bg="#F0F0F0").pack(pady=4)

    #Modulo derecho
    tk.Label(frame_der, text="Ventas realizadas", font=("Arial", 14, "bold"), bg="#F0F0F0").pack(pady=8)

    # Frame para centrar el buscador
    frame_buscar = tk.Frame(frame_der, bg="#F0F0F0")
    frame_buscar.pack(pady=5)

  # Contenedor centrado
    frame_busqueda_venta = tk.Frame(frame_der, bg="#F0F0F0")
    frame_busqueda_venta.pack(pady=10, padx=50)  # padx para "pared" a la izquierda

# Label Buscar
    tk.Label(frame_busqueda_venta, text="Buscar:", font=("Arial", 12), bg="#F0F0F0").pack(side="left", padx=(0,5))

# Entry Buscar
    entry_buscar = tk.Entry(frame_busqueda_venta, width=40, justify="center")
    entry_buscar.pack(side="left", padx=5)
    entry_buscar.bind("<KeyRelease>", lambda e: actualizar_tabla(entry_buscar.get()))

# Botón Anular Venta
    ttk.Button(frame_busqueda_venta, text="Anular Venta", padding=5, bootstyle="danger-outline",
           command=anular_venta).pack(side="left", padx=(5,0))

    # Tabla de ventas
    frame_tabla = tk.Frame(frame_der, bg="#F0F0F0")
    frame_tabla.pack(fill="both", expand=True, padx=5, pady=5)
    columnas = ("Producto", "Cliente", "Cantidad", "Total")
    tabla = ttk.Treeview(frame_tabla, columns=columnas, show="headings", height=16)
    for col in columnas:
        tabla.heading(col, text=col)
        tabla.column(col, width=130, anchor="center")
    tabla.pack(fill="both", expand=True)

    

    # activar autocompletado
    autocompletar(entry_cliente, [n for _, n, *_ in leer_hoja("Clientes")])
    autocompletar(entry_producto, [n for _, n, *_ in leer_hoja("Inventario")])

    actualizar_tabla()
    actualizar_tabla_temporal()
