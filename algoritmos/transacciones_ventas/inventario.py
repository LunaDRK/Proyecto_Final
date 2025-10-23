import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
import ttkbootstrap as ttk
from ttkbootstrap.constants import *

def limpiar_ventana(ventana):
    for widget in ventana.winfo_children():
        widget.destroy()

def mostrar_inventario(ventana, volver_callback):
    limpiar_ventana(ventana)

    # Botón de regresar
    # Botón de regresar
    btn_volver = ttk.Button(ventana, text="← Regresar", width=15, padding=10, bootstyle="danger-outline",
                           command=volver_callback)
    btn_volver.pack(side="top", anchor="nw", padx=15, pady=15)

    #Base de datos y tabla de inventario
    ARCHIVO = "algoritmos/transacciones_ventas/proyecto.xlsx"
    HOJA = "Inventario"

    # Crear hoja si no existen 
    def crear_excel_y_hoja():
        if not os.path.exists(ARCHIVO):
            wb = openpyxl.Workbook()
            if HOJA in wb.sheetnames:
                hoja = wb[HOJA]
            else:
                hoja = wb.active
                hoja.title = HOJA
            hoja["A1"] = "Código"
            hoja["B1"] = "Nombre"
            hoja["C1"] = "Existencia"
            hoja["D1"] = "Proveedor"
            hoja["E1"] = "Precio"
            wb.save(ARCHIVO)
            wb.close()
        else:
            wb = openpyxl.load_workbook(ARCHIVO)
            if HOJA not in wb.sheetnames:
                hoja = wb.create_sheet(HOJA)
                hoja["A1"] = "Código"
                hoja["B1"] = "Nombre"
                hoja["C1"] = "Existencia"
                hoja["D1"] = "Proveedor"
                hoja["E1"] = "Precio"
                wb.save(ARCHIVO)
            wb.close()

    # Leer productos 
    def leer_productos():
        wb = openpyxl.load_workbook(ARCHIVO)
        hoja = wb[HOJA]
        datos = []
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            datos.append(fila)
        wb.close()
        return datos

    # Escribir nuevo producto 
    def escribir_producto(codigo, nombre, existencia, proveedor, precio):
        if not codigo or not nombre or not existencia or not proveedor or not precio:
            messagebox.showwarning("Campos vacíos", "Debes llenar todos los campos.")
            return

        try:
            existencia = int(existencia)
            precio = float(precio)
        except ValueError:
            messagebox.showerror("Error", "Existencia debe ser entero y Precio número (ej: 12.50).")
            return

        # Append en la hoja correcta sin tocar otras hojas
        wb = openpyxl.load_workbook(ARCHIVO)
        hoja = wb[HOJA]
        hoja.append([codigo, nombre, existencia, proveedor, precio])
        wb.save(ARCHIVO)
        wb.close()
        messagebox.showinfo("Éxito", "Producto agregado correctamente.")
        actualizar_tabla()
        limpiar_campos()

    # Actualizar tabla en la interfaz
    def actualizar_tabla():
        for fila in tabla.get_children():
            tabla.delete(fila)
        for producto in leer_productos():
            tabla.insert("", "end", values=producto)

    # Buscar productos 
    def buscar_producto():
        texto = entry_buscar.get().strip().lower()
        if not texto:
            messagebox.showinfo("Buscar", "Ingrese algo para buscar.")
            return

        resultados = []
        for producto in leer_productos():
            if any(texto in str(campo).lower() for campo in producto if campo is not None):
                resultados.append(producto)

        for fila in tabla.get_children():
            tabla.delete(fila)

        if resultados:
            for producto in resultados:
                tabla.insert("", "end", values=producto)
        else:
            messagebox.showinfo("Sin resultados", "No se encontraron coincidencias.")

    def limpiar_campos():
        entry_codigo.delete(0, tk.END)
        entry_nombre.delete(0, tk.END)
        entry_existencia.delete(0, tk.END)
        entry_proveedor.delete(0, tk.END)
        entry_precio.delete(0, tk.END)

    def editar_productos():
        producto_seleccionado = tabla.focus()
        if not producto_seleccionado:
            messagebox.showwarning("Editar", "Selecciona dando un click sobre el producto para editar")
            return
        
        valores_originales = tabla.item(producto_seleccionado, "values")
        codigo_original, nombre_original, existencia_original, proveedor_original, precio_original = valores_originales

        #crea una ventana emergente para un form de editar
        ventana_editar = tk.Toplevel(ventana)
        ventana_editar.title("Editar Producto")
        ventana_editar.geometry("400x550")
        ventana_editar.configure(bg="#F0F0F0")

        tk.Label(ventana_editar, text="Código:", font=("Arial", 12), bg="#F0F0F0").grid(row=0, column=0, padx=10, pady=10)
        entry_codigo_editar = tk.Entry(ventana_editar, width=30)
        entry_codigo_editar.grid(row=0, column=1, padx=10, pady=10)
        entry_codigo_editar.insert(0, codigo_original)

        tk.Label(ventana_editar, text="Nombre:", font=("Arial", 12), bg="#F0F0F0").grid(row=1, column=0, padx=10, pady=10)
        entry_nombre_editar = tk.Entry(ventana_editar, width=30)
        entry_nombre_editar.grid(row=1, column=1, padx=10, pady=10)
        entry_nombre_editar.insert(0, nombre_original)

        tk.Label(ventana_editar, text="Existencia:", font=("Arial", 12), bg="#F0F0F0").grid(row=2, column=0, padx=10, pady=10)
        entry_existencia_editar = tk.Entry(ventana_editar, width=30)
        entry_existencia_editar.grid(row=2, column=1, padx=10, pady=10)
        entry_existencia_editar.insert(0, existencia_original)

        tk.Label(ventana_editar, text="Proveedor:", font=("Arial", 12), bg="#F0F0F0").grid(row=3, column=0, padx=10, pady=10)
        entry_proveedor_editar = tk.Entry(ventana_editar, width=30)
        entry_proveedor_editar.grid(row=3, column=1, padx=10, pady=10)
        entry_proveedor_editar.insert(0, proveedor_original)

        tk.Label(ventana_editar, text="Precio:", font=("Arial", 12), bg="#F0F0F0").grid(row=4, column=0, padx=10, pady=10)
        entry_precio_editar = tk.Entry(ventana_editar, width=30)
        entry_precio_editar.grid(row=4, column=1, padx=10, pady=10)
        entry_precio_editar.insert(0, precio_original)

        def guardar_cambios():
            codigo_editado = entry_codigo_editar.get().strip()
            nombre_editado = entry_nombre_editar.get().strip()
            existencia_editada = entry_existencia_editar.get().strip()
            proveedror_editada = entry_proveedor_editar.get().strip()
            precio_editada = entry_precio_editar.get().strip()

            if not codigo_editado or not nombre_editado or not existencia_editada or not proveedror_editada or not precio_editada:
                messagebox.showwarning("Campos vacíos", "Debes llenar al menos un campo.")
                return
            
             # Actualizar el excel
            wb = openpyxl.load_workbook(ARCHIVO)
            hoja = wb[HOJA]
            for fila in hoja.iter_rows(min_row=2):
                if str(fila[0].value) == str(codigo_original):
                    fila[0].value = codigo_editado
                    fila[1].value = nombre_editado
                    fila[2].value = existencia_editada
                    fila[3].value = proveedror_editada
                    fila[4].value = precio_editada
                    break

            wb.save(ARCHIVO)
            wb.close()

            # Actualiza solo la fila que se editó y muestra mensaje que se editó correctamente
            tabla.item(producto_seleccionado, values=(codigo_editado, nombre_editado, existencia_editada, proveedror_editada, precio_editada))
            messagebox.showinfo("Editado", "Información de producto actualizada correctamente")
            ventana_editar.destroy()  # Cierra la ventana de edición

        tk.Button(ventana_editar, text="Guardar Cambios", bg="#FFC107", fg="black",
                        font=("Arial", 12, "bold"), command=guardar_cambios).grid(row=5, column=0, columnspan=2, pady=20)
        
    def eliminar_producto():
        producto_seleccionado = tabla.focus()
        if not producto_seleccionado:
            messagebox.showwarning("Eliminar", "Selecciona un producto para eliminar")
            return

        valores = tabla.item(producto_seleccionado, "values")
        codigo_producto = valores[0]
        nombre_producto = valores[1]

        confirmar = messagebox.askyesno("Confirmar Eliminación", f"¿Estás seguro de eliminar el producto {nombre_producto} con código {codigo_producto}?")
        if not confirmar:
            return
        
        # Elimina del excel
        wb = openpyxl.load_workbook(ARCHIVO)
        hoja = wb[HOJA]
        for fila in hoja.iter_rows(min_row=2):
            if str(fila[0].value) == str(codigo_producto):
                hoja.delete_rows(fila[0].row, 1)
                break

        # Eliminar todas las ventas relacionadas con el producto
        if "Ventas" in wb.sheetnames:
            hoja_ventas = wb["Ventas"]
            filas_a_eliminar = []
            for fila in range(2, hoja_ventas.max_row + 1):
                if str(hoja_ventas.cell(fila, 1).value) == str(codigo_producto):
                    filas_a_eliminar.append(fila)
            for fila in reversed(filas_a_eliminar):
                hoja_ventas.delete_rows(fila, 1)

        wb.save(ARCHIVO)
        wb.close()

        # Elimina de la tabla y muestra mensaje que se eliminó correctamente
        tabla.delete(producto_seleccionado)
        messagebox.showinfo("Eliminado", "Producto eliminado correctamente.")

    # Inicializador
    crear_excel_y_hoja()

    titulo = tk.Label(ventana, text="Control de Inventario", font=("Arial", 18, "bold"), bg="#F0F0F0")
    titulo.pack(pady=15, side='top')

    # Frame del formulario
    frame_form = tk.Frame(ventana, bg="#F0F0F0")
    frame_form.pack(pady=10)

    tk.Label(frame_form, text="Código:", font=("Arial", 12), bg="#F0F0F0").grid(row=0, column=0, padx=10, pady=5)
    entry_codigo = tk.Entry(frame_form, width=30)
    entry_codigo.grid(row=0, column=1, padx=10, pady=5)

    tk.Label(frame_form, text="Nombre:", font=("Arial", 12), bg="#F0F0F0").grid(row=1, column=0, padx=10, pady=5)
    entry_nombre = tk.Entry(frame_form, width=30)
    entry_nombre.grid(row=1, column=1, padx=10, pady=5)

    tk.Label(frame_form, text="Existencia:", font=("Arial", 12), bg="#F0F0F0").grid(row=2, column=0, padx=10, pady=5)
    entry_existencia = tk.Entry(frame_form, width=30)
    entry_existencia.grid(row=2, column=1, padx=10, pady=5)

    tk.Label(frame_form, text="Proveedor:", font=("Arial", 12), bg="#F0F0F0").grid(row=3, column=0, padx=10, pady=5)
    entry_proveedor = tk.Entry(frame_form, width=30)
    entry_proveedor.grid(row=3, column=1, padx=10, pady=5)

    tk.Label(frame_form, text="Precio:", font=("Arial", 12), bg="#F0F0F0").grid(row=4, column=0, padx=10, pady=5)
    entry_precio = tk.Entry(frame_form, width=30)
    entry_precio.grid(row=4, column=1, padx=10, pady=5)

    btn_agregar = ttk.Button(ventana, text="Agregar Producto", padding=5, bootstyle="success",
                             width=20,
                            command=lambda: escribir_producto(entry_codigo.get(), entry_nombre.get(),
                                                              entry_existencia.get(), entry_proveedor.get(),
                                                              entry_precio.get()))
    btn_agregar.pack(pady=10)

    # Frame de búsqueda
    frame_buscar = tk.Frame(ventana, bg="#F0F0F0")
    frame_buscar.pack(pady=10)

    tk.Label(frame_buscar, text="Buscar:", font=("Arial", 12), bg="#F0F0F0").grid(row=0, column=0, padx=10)
    entry_buscar = tk.Entry(frame_buscar, width=40)
    entry_buscar.grid(row=0, column=1, padx=10)
    btn_buscar = ttk.Button(frame_buscar, text="Buscar", padding=5, bootstyle="warning-outline",
                        width=10, command=buscar_producto)
    btn_buscar.grid(row=0, column=2, padx=10)

    btn_todos = ttk.Button(frame_buscar, text="Mostrar Todos", padding=5, bootstyle="info-outline",
                        width=15, command=actualizar_tabla)
    btn_todos.grid(row=0, column=3, padx=10)

    #btn de editar y eliminar
    frame_botones = tk.Frame(ventana, bg="#F0F0F0")
    frame_botones.pack(pady=10)

    ttk.Button(frame_buscar, text="Editar", padding=5, bootstyle="success-outline", width=10,
                    command=editar_productos).grid(row=0, column=4, padx=5)

    ttk.Button(frame_buscar, text="Eliminar", padding=5, bootstyle="danger-outline", width=10,
             command=eliminar_producto).grid(row=0, column=5, padx=5)

    # Tabla
    frame_tabla = tk.Frame(ventana)
    frame_tabla.pack(pady=10)

    columnas = ("Código", "Nombre", "Existencia", "Proveedor", "Precio")
    tabla = ttk.Treeview(frame_tabla, columns=columnas, show="headings", height=12)
    for col in columnas:
        tabla.heading(col, text=col)
        tabla.column(col, width=150, anchor="center")
    tabla.pack()

    actualizar_tabla()